"""
Converted from your Power Query (M) to Python (pandas), without lookups.xlsx.

This script:
- Finds the latest VW_ONEMI_ESTATE_MANAGEMENT_*.xlsx in a folder
- Loads Sheet1
- Filters REMOVED_FLAG is null
- Dedupes by PLANNER_UNIQUE_IDENTIFIER
- Renames OneMI *_TODAY columns
- Replaces empty/null cells in specific columns with "@_EMPTY"
- Fills invalid DISPOSITION_TIMELINE_STEP_1 with 1900-01-01
- Loads lookup tables exactly like your PQ:
    - Settings.xlsx (Excel Tables)
    - Schedules.xlsx (Excel Tables)
    - NAR_ReportBaseLine.xlsx (Sheet)
- Left-joins all lookup tables
- Derives Location_Today + Location_Calculated
- Derives Technology_Today + Technology_Calculated
- Fills UNKNOWN location/technology from manual settings
- Builds S2T2T / S2T-1 / S2T-2
- Adds chase decision columns
- Merges schedules, data residency, shared/dedicated tagging, DC-name based location
- Drops many columns
- Reorders columns
"""

from __future__ import annotations

import os
import glob
from pathlib import Path
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook


# ----------------------------
# CONFIG (EDIT THESE)
# ----------------------------
FOLDER_WITH_ONEMI_EXPORTS = r"C:\Users\arwin\OneDrive\Desktop\AID\input"
FILE_PREFIX = "VW_ONEMI_ESTATE_MANAGEMENT_"
SHEET_NAME_MAIN = "Sheet1"

# These correspond to your Power Query sources
SETTINGS_XLSX_PATH = r"C:\Users\auerleo\Deutsche Bank AG\Technology Transformation And Integration - 08 App Infra Disposition\90 - MasterMaker\Master_Maker_2.1\Settings.xlsx"
SCHEDULES_XLSX_PATH = r"C:\Users\auerleo\Deutsche Bank AG\Technology Transformation And Integration - 08 App Infra Disposition\90 - MasterMaker\Master_Maker_2.1\Schedules.xlsx"
NAR_BASELINE_XLSX_PATH = r"C:\Users\auerleo\Deutsche Bank AG\Technology Transformation And Integration - 08 App Infra Disposition\90 - MasterMaker\Master_Maker_2.1\NAR_ReportBaseLine.xlsx"
NAR_BASELINE_SHEET_NAME = "NAR_ReportBaseLine"

# Output
OUTPUT_PATH = r"C:\Users\arwin\OneDrive\Desktop\AID\estate_management_output.xlsx"

EMPTY_TOKEN = "@_EMPTY"


# ----------------------------
# Helpers
# ----------------------------
def latest_matching_excel(folder: str, prefix: str) -> str:
    pattern = str(Path(folder) / f"{prefix}*.xlsx")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"No files found matching: {pattern}")
    return max(files, key=os.path.getmtime)


def ensure_exists(path: str, label: str) -> None:
    if not path or not Path(path).exists():
        raise FileNotFoundError(f"{label} not found: {path}")


def _a1_to_idx(a1: str) -> Tuple[int, int]:
    # "A1" -> (row=1, col=1)
    col = 0
    row_str = ""
    for ch in a1:
        if ch.isalpha():
            col = col * 26 + (ord(ch.upper()) - ord("A") + 1)
        else:
            row_str += ch
    return int(row_str), int(col)


def _range_to_bounds(rng: str) -> Tuple[int, int, int, int]:
    # "A1:D10" -> (min_row, min_col, max_row, max_col)
    left, right = rng.split(":")
    r1, c1 = _a1_to_idx(left)
    r2, c2 = _a1_to_idx(right)
    return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)


def read_excel_table(workbook_path: str, table_name: str) -> pd.DataFrame:
    """
    Reads an Excel *Table object* (not a sheet) by table name, like Power Query's:
      Source{[Item="tabCIsettings",Kind="Table"]}[Data]
    """
    ensure_exists(workbook_path, "Workbook")
    wb = load_workbook(workbook_path, data_only=True, read_only=True)

    for ws in wb.worksheets:
        # openpyxl stores tables in ws.tables
        if table_name in ws.tables:
            tbl = ws.tables[table_name]
            min_row, min_col, max_row, max_col = _range_to_bounds(tbl.ref)

            data = []
            for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col, values_only=True):
                data.append(list(row))

            if not data:
                return pd.DataFrame()

            headers = [str(h).strip() if h is not None else "" for h in data[0]]
            rows = data[1:]
            df = pd.DataFrame(rows, columns=headers)
            return df

    raise KeyError(f"Excel table '{table_name}' not found in workbook: {workbook_path}")


def fill_empty_cells(df: pd.DataFrame, cols: List[str], token: str) -> pd.DataFrame:
    for c in cols:
        if c not in df.columns:
            continue
        df[c] = df[c].astype("string").fillna("").map(lambda v: token if str(v).strip() == "" else v)
    return df


def safe_contains(val: object, needle: str) -> bool:
    if val is None:
        return False
    if isinstance(val, float) and np.isnan(val):
        return False
    return needle in str(val)


def left_merge(df: pd.DataFrame, right: pd.DataFrame, left_on: List[str], right_on: List[str]) -> pd.DataFrame:
    return df.merge(right, how="left", left_on=left_on, right_on=right_on)


def ensure_col(df: pd.DataFrame, col: str, default) -> None:
    if col not in df.columns:
        df[col] = default


# ----------------------------
# Load lookups exactly like Power Query
# ----------------------------
def load_tabCIsettings() -> pd.DataFrame:
    df = read_excel_table(SETTINGS_XLSX_PATH, "tabCIsettings")
    # PQ:
    #  - rename "CI Summary" -> CI_Summary_File
    #  - new "CI Summary": if contains "No response required" then "No response required" else file value
    if "CI Summary" in df.columns:
        df = df.rename(columns={"CI Summary": "CI_Summary_File"})
        df["CI Summary"] = df["CI_Summary_File"].astype("string").fillna("").apply(
            lambda x: "No response required" if "No response required" in str(x) else str(x)
        )
        df = df.drop(columns=["CI_Summary_File"])
    return df


def load_tabStep0settings() -> pd.DataFrame:
    # PQ reads table "tabStepsettings"
    df = read_excel_table(SETTINGS_XLSX_PATH, "tabStepsettings")
    # Keep the columns as-is; joins expect Technology Today vs Technology_Today mismatch handled later
    return df


def load_tabStep1settings() -> pd.DataFrame:
    return read_excel_table(SETTINGS_XLSX_PATH, "tabStep1settings")


def load_tabStep2settings() -> pd.DataFrame:
    return read_excel_table(SETTINGS_XLSX_PATH, "tabStep2settings")


def load_tabS2T1settings() -> pd.DataFrame:
    df = read_excel_table(SETTINGS_XLSX_PATH, "tabS2T1settings")
    # Your main PQ expects: "S2T1-CIO to Dispo Chase" (but the snippet shows "Dispo Chase")
    # In the saved versions you had both; handle safely:
    if "Dispo Chase" in df.columns and "S2T1-CIO to Dispo Chase" not in df.columns:
        df = df.rename(columns={"Dispo Chase": "S2T1-CIO to Dispo Chase"})
    return df


def load_tabS2T2settings() -> pd.DataFrame:
    df = read_excel_table(SETTINGS_XLSX_PATH, "tabS2T2settings")
    if "Dispo Chase" in df.columns and "S2T2-CIO to Dispo Chase" not in df.columns:
        df = df.rename(columns={"Dispo Chase": "S2T2-CIO to Dispo Chase"})
    return df


def load_tabS2T2Tsettings() -> pd.DataFrame:
    df = read_excel_table(SETTINGS_XLSX_PATH, "tabS2T2Tsettings")
    # Main python expects "CIO Chase YN" not "CIO Decision" in some earlier versions
    if "CIO Decision" in df.columns and "CIO Chase YN" not in df.columns:
        df = df.rename(columns={"CIO Decision": "CIO Chase YN"})
    return df


def load_NAR_ReportBaseLine() -> pd.DataFrame:
    ensure_exists(NAR_BASELINE_XLSX_PATH, "NAR baseline workbook")
    df = pd.read_excel(NAR_BASELINE_XLSX_PATH, sheet_name=NAR_BASELINE_SHEET_NAME, engine="openpyxl")
    return df


def load_tabPlatformNARs() -> pd.DataFrame:
    df = read_excel_table(SETTINGS_XLSX_PATH, "tabPlatformNARs")
    # PQ rename ReturnValue -> Platform_Provider, replace yes -> Yes
    if "ReturnValue" in df.columns:
        df = df.rename(columns={"ReturnValue": "Platform_Provider"})
    if "Platform_Provider" in df.columns:
        df["Platform_Provider"] = df["Platform_Provider"].astype("string").fillna("").str.replace("yes", "Yes", regex=False)
    return df


def load_tabSchedule_V2V() -> pd.DataFrame:
    df = read_excel_table(SCHEDULES_XLSX_PATH, "tabSchedule_V2V")
    # PQ: replace null Full_Server_Name with "@_Empty"
    if "Full_Server_Name" in df.columns:
        df["Full_Server_Name"] = df["Full_Server_Name"].astype("string").fillna("@_Empty")
    # PQ: replace "United Kingdom" -> "UK" in Country
    if "Country" in df.columns:
        df["Country"] = df["Country"].astype("string").fillna("").str.replace("United Kingdom", "UK", regex=False)

    # PQ adds V2V-Scope from "Migration wave"
    # (your stored V2V query also renames wave start/end columns; we keep any columns that exist)
    if "Migration wave" in df.columns:
        def v2v_scope(row) -> str:
            wave = str(row.get("Migration wave", "") or "")
            country = str(row.get("Country", "") or "")
            if "subnet not found" in wave:
                return f"{country}-TBC"
            if "subnet without virtual instance" in wave:
                return f"{country}-TBC"
            return wave
        df["V2V-Scope"] = df.apply(v2v_scope, axis=1)

    return df


def load_tabSchedule_P2V() -> pd.DataFrame:
    return read_excel_table(SCHEDULES_XLSX_PATH, "tabSchedule_P2V")


def load_tabSchedule_P2P() -> pd.DataFrame:
    return read_excel_table(SCHEDULES_XLSX_PATH, "tabSchedule_P2P")


def load_tabData_Residency() -> pd.DataFrame:
    df = read_excel_table(SETTINGS_XLSX_PATH, "Data_Residency")
    # PQ renames "Data Residency" -> "Data_Residency"
    if "Data Residency" in df.columns:
        df = df.rename(columns={"Data Residency": "Data_Residency"})
    return df


def load_tabUnderpinningDBServer() -> pd.DataFrame:
    df = read_excel_table(SETTINGS_XLSX_PATH, "Database_server_Name")
    # PQ distinct by SERVER_NAME (we do that)
    if "SERVER_NAME" in df.columns:
        df = df.drop_duplicates(subset=["SERVER_NAME"], keep="first").copy()
    # Some versions also have Underpinning_Server_CIs; keep whatever exists
    return df


def load_Location_Today2() -> pd.DataFrame:
    return read_excel_table(SETTINGS_XLSX_PATH, "Location_Today")


def load_tabEAP_Grid_Consumers() -> pd.DataFrame:
    df = read_excel_table(SETTINGS_XLSX_PATH, "tabEAP_Grid_Consumers")
    # PQ renames Status -> EAP/Grid Consumer
    if "Status" in df.columns:
        df = df.rename(columns={"Status": "EAP/Grid Consumer"})
    return df


# NOTE: You didn’t paste M for Server_Shared-Dedicate_Tagging in this message,
# but your main query uses it. You already had it earlier in the conversation memory.
# So we load it as a function expecting it to be available as a TABLE in Settings.xlsx
# If your real source is different, update this function accordingly.
def load_Server_Shared_Dedicate_Tagging() -> pd.DataFrame:
    # If you already have it as an Excel Table named exactly like PQ query name:
    # try to read it; otherwise you must point to its real workbook source.
    try:
        return read_excel_table(SETTINGS_XLSX_PATH, "Server_Shared-Dedicate_Tagging")
    except Exception:
        # Fallback name used in some workbooks
        return read_excel_table(SETTINGS_XLSX_PATH, "Server_Shared-Dedicate_Tagging".replace("-", "_"))


# ----------------------------
# Core transformations
# ----------------------------
def compute_location_today(r: pd.Series) -> str:
    country = str(r.get("COUNTRY", "") or "")
    building = str(r.get("BUILDING", "") or "")
    billable = r.get("HP_P4_P7_BILLABLE", None)

    # Hub-Locations
    if (
        (country == "GERMANY" and building in {"KRUPPSTRASSE 121 - 127 (DCB)", "LAERCHENSTRASSE 110 (DCN)"})
        or ((country in {"UK", "UNITED KINGDOM"}) and building in {"CROYDON DATA CENTRE", "WATFORD DATA CENTRE"})
        or ((country in {"USA", "UNITED STATES OF AMERICA"}) and building in {"3 CORPORATE PLACE", "USZPK"})
        or (country == "SINGAPORE" and building in {"DSJ", "38 KIM CHUAN", "CAPITALAND 9 TAI SENG DRIVE"})
    ):
        return "HUB"

    # Blaupause
    if country == "GERMANY" and building in {"GABLONZER STRASSE 34 (DCO)", "BISMARCKSTRASSE 2 (DCS)"}:
        return "Blaupause"

    # GCP Cloud Locations
    if building in {
        "EUROPE-WEST3-ZONE-A", "EUROPE-WEST3-ZONE-B", "EUROPE-WEST3-ZONE-C",
        "EUROPE-WEST2-ZONE-A", "EUROPE-WEST2-ZONE-B", "EUROPE-WEST2-ZONE-C",
        "US-EAST4-ZONE-A", "US-EAST4-ZONE-B", "US-EAST4-ZONE-C",
    }:
        return "GCP-Cloud"

    # Non-Hub Locations - non-billable
    building_ok = building not in {"", "None"} and building is not None
    if billable == "N" and building_ok:
        return "Non-Hub (non-billable)"

    # Non-Hub Locations - billable
    if billable == "Y" and building not in {"", EMPTY_TOKEN}:
        return "Non-Hub (billable)"

    # Non-Hub Locations - billable tbc
    if billable in [None, ""] and building not in {"", EMPTY_TOKEN}:
        return "Non-Hub (billable tbc)"

    return "UNKNOWN"


def compute_technology_today(r: pd.Series) -> str:
    ci_cat = str(r.get("CI_CATEGORY", "") or "")
    is_actual_db = str(r.get("IS_ACTUAL_DATABASE_CI", "") or "")
    building = str(r.get("BUILDING", "") or "")
    service = str(r.get("SERVICE", "") or "")
    db_type = str(r.get("DATABASE_TYPE", "") or "")
    db_ver = str(r.get("DATABASE_VERSION", "") or "")
    underpin = str(r.get("Underpinning_Server_CIs", "") or "")
    instance_name = str(r.get("INSTANCE_NAME", "") or "")
    service_offering = str(r.get("SERVICE_OFFERING", "") or "")
    portfolio_nar = str(r.get("PRODUCT_PORTFOLIO_NAR", "") or "")
    server_type = str(r.get("SERVER_TYPE", "") or "")
    host_type = str(r.get("HOST_TYPE", "") or "")

    if ci_cat == "DATABASE" and is_actual_db == "N":
        return "No response required (No real Oracle DB)"

    if "ZONE" in building:
        return "No response required (GCP CIs)"

    if ci_cat == "SERVER" and underpin == "Y":
        return "No response required (Server underpinning CI Database)"

    if service in {"DAP", "dWeb", "Fabric"}:
        return "PaaS"

    if ci_cat == "DATABASE" and db_type == "MICROSOFT" and "MICROSOFT SQL SERVER" in db_ver:
        return "SQL"

    if ci_cat == "DATABASE" and db_type == "SYBASE":
        return "Sybase"

    if ci_cat == "DATABASE" and db_type == "SAP" and db_ver in {"SAP HANA ENTERPRISE EDITION", "SAP REPLICATION SERVER 16.0"}:
        return "SAP"

    if ci_cat == "DATABASE" and db_ver in {"ORACLE DATABASE 19", "ORACLE RAC 19", "ORACLE GRID 19"}:
        return "Oracle 19"

    if ci_cat == "DATABASE" and db_ver in {
        "ORACLE DATABASE 10.2.0.5.0",
        "ORACLE DATABASE 11.2.0.3.0",
        "ORACLE DATABASE 11.2.0.4.0",
        "ORACLE DATABASE 12.1.0.2.0",
        "ORACLE DATABASE 12.2.0.1",
        "ORACLE RAC 11.2.0.4.0",
        "ORACLE GRID INFRASTRUCTURE 12.1.0.2.0",
        "ORACLE DATABASE 18",
    }:
        return "Oracle Legacy"

    if ci_cat == "SERVER" and instance_name in {
        "DAP-GRLOBAL", "dWEB-GRLOBAL", "FABRIC-GLOBAL", "EAP-tools",
        "EAP-UK-Big Data Platform", "EAP-DE-Big Data Platform",
    }:
        return "Hosting - PaaS"

    if ci_cat == "SERVER" and service_offering in {
        "dCloud Database Oracle - Premium",
        "dCloud Database Oracle - Shared",
        "ODA - OFBA|ODA - OFBB",
        "ODA - OFBA",
        "ODA - OFBB",
    }:
        return "Hosting-Oracle"

    if ci_cat == "SERVER" and portfolio_nar == "Y":
        return "Hosting"

    if ci_cat == "SERVER" and service_offering == "EXADATA SHARED SERVICE":
        return "Standalone Exa"

    if ci_cat == "SERVER" and service_offering in {"HADOOP SHARED SERVICE", "HADOOP SHARED SERVICE|Harvested Grid SO"}:
        return "Hadoop"

    if ci_cat == "SERVER" and (service == "GRID" or service_offering in {"Native Grid SO", "Harvested Grid SO"}):
        return "GRID Compute"

    if ci_cat == "SERVER" and server_type == "AIX":
        return "Legacy Compute - AIX"

    if ci_cat == "SERVER" and server_type == "SPARC":
        return "Legacy Compute - SPARC"

    if ci_cat == "SERVER" and service == "VHS":
        return "VHS"

    if ci_cat == "SERVER" and server_type == "X86_VIRTUAL" and service != "VHS":
        return "x86 Virtual"

    if ci_cat == "SERVER" and host_type == "VIRTUAL":
        return "x86 Virtual"

    if ci_cat == "SERVER" and server_type == "X86_PHYSICAL" and portfolio_nar != "Y":
        return "x86 Physical"

    if ci_cat == "SERVER" and host_type == "PHYSICAL":
        return "x86 Physical"

    return "UNKNOWN"


def compute_v2v_sub_scope(r: pd.Series) -> str:
    try:
        if (
            str(r.get("HOST_TYPE", "")) == "VIRTUAL"
            and str(r.get("HUB_LOCATION", "")) == "Y"
            and (safe_contains(r.get("MODEL"), "PROLIANT DL6") or safe_contains(r.get("MODEL"), "PROLIANT BL4"))
        ):
            return "VHS on C7000"
        if (
            str(r.get("HOST_TYPE", "")) == "VIRTUAL"
            and str(r.get("HUB_LOCATION", "")) == "Y"
            and safe_contains(r.get("MODEL"), "SY480")
        ):
            return "VHS on Synergy"
        return "0_EMPTY"
    except Exception:
        return "0_EMPTY"


def compute_vendor_today(r: pd.Series) -> str:
    try:
        ci_cat = str(r.get("CI_CATEGORY", "") or "")
        model = r.get("MODEL", None)
        os_ = r.get("OPERATING_SYSTEM", None)

        if ci_cat == "DATABASE" and (safe_contains(model, "ORACLE") or safe_contains(model, "Oracle")):
            return "Oracle"
        if ci_cat == "DATABASE" and safe_contains(model, "MICROSOFT"):
            return "Microsoft"
        if ci_cat == "SERVER" and (safe_contains(os_, "ORACLE") or safe_contains(os_, "Oracle")):
            return "Oracle"
        if ci_cat == "SERVER" and (safe_contains(os_, "WINDOWS SERVER") or safe_contains(os_, "Windows Server")):
            return "Microsoft"
        if ci_cat == "SERVER" and (safe_contains(os_, "REDHAT") or safe_contains(os_, "Red Hat")):
            return "RedHat"
        if ci_cat == "SERVER" and safe_contains(os_, "SLES"):
            return "SUSE Linux"
        if ci_cat == "SERVER" and safe_contains(os_, "VMWARE ESXI"):
            return "VMware"
        return "0_EMPTY"
    except Exception:
        return "0_EMPTY"


# ----------------------------
# Main pipeline
# ----------------------------
def run_pipeline() -> pd.DataFrame:
    latest_file = latest_matching_excel(FOLDER_WITH_ONEMI_EXPORTS, FILE_PREFIX)
    df = pd.read_excel(latest_file, sheet_name=SHEET_NAME_MAIN, engine="openpyxl")

    # Step 7: Active filter
    if "REMOVED_FLAG" in df.columns:
        df = df[df["REMOVED_FLAG"].isna()].copy()

    # Step 8: Dedupe
    if "PLANNER_UNIQUE_IDENTIFIER" in df.columns:
        df = df.drop_duplicates(subset=["PLANNER_UNIQUE_IDENTIFIER"], keep="first").copy()

    # Rename OneMI columns
    df = df.rename(columns={
        "LOCATION_TODAY": "OneMI_LOCATION_TODAY",
        "TECHNOLOGY_TODAY": "OneMI_TECHNOLOGY_TODAY",
        "VENDOR_TODAY": "OneMI_VENDOR_TODAY",
    })

    # Replace EMPTY Cells
    df = fill_empty_cells(
        df,
        cols=[
            "DISPOSITION_OPTION_STEP_1",
            "DISPOSITION_OPTION_STEP_2",
            "DISPOSITION_TIMELINE_STEP_2",
            "COUNTRY",
            "BUILDING",
            "HP_DC_NAME",
            "DATABASE_VERSION",
            "DISPOSITION_COMMENTS_STEP_1",
            "DISPOSITION_COMMENTS_STEP_2",
        ],
        token=EMPTY_TOKEN,
    )

    # Replace Errors for DISPOSITION_TIMELINE_STEP_1
    if "DISPOSITION_TIMELINE_STEP_1" in df.columns:
        dt = pd.to_datetime(df["DISPOSITION_TIMELINE_STEP_1"], errors="coerce")
        df["DISPOSITION_TIMELINE_STEP_1"] = dt.fillna(pd.Timestamp("1900-01-01")).dt.date

    # ----------------------------
    # Load lookups from real sources (like PQ)
    # ----------------------------
    tabUnderpinningDBServer = load_tabUnderpinningDBServer()
    tabCIsettings = load_tabCIsettings()
    tabStep0settings = load_tabStep0settings()
    tabStep1settings = load_tabStep1settings()
    tabStep2settings = load_tabStep2settings()
    tabS2T1settings = load_tabS2T1settings()
    tabS2T2settings = load_tabS2T2settings()
    tabS2T2Tsettings = load_tabS2T2Tsettings()
    NAR_ReportBaseLine = load_NAR_ReportBaseLine()
    tabEAP_Grid_Consumers = load_tabEAP_Grid_Consumers()
    tabPlatformNARs = load_tabPlatformNARs()
    tabSchedule_V2V = load_tabSchedule_V2V()
    tabSchedule_P2V = load_tabSchedule_P2V()
    tabSchedule_P2P = load_tabSchedule_P2P()
    tabData_Residency = load_tabData_Residency()
    Server_Shared_Dedicate_Tagging = load_Server_Shared_Dedicate_Tagging()
    Location_Today2 = load_Location_Today2()

    # ----------------------------
    # Merge underpinning DB server
    # PQ expands {"Flag","Underpinning_Server_CIs"} if present
    # ----------------------------
    df = left_merge(df, tabUnderpinningDBServer, ["SERVER_NAME"], ["SERVER_NAME"])

    # Location_Today + calculated flag
    ensure_col(df, "COUNTRY", None)
    ensure_col(df, "BUILDING", None)
    ensure_col(df, "HP_P4_P7_BILLABLE", None)
    df["Location_Today"] = df.apply(compute_location_today, axis=1)
    df["Location_Calculated"] = np.where(df["Location_Today"] != "UNKNOWN", "Calculated", "Manual")

    # HUB_LOCATION for later V2V sub-scope logic
    df["HUB_LOCATION"] = np.where(df["Location_Today"] == "HUB", "Y", "N")

    # Technology_Today + calculated flag
    df["Technology_Today"] = df.apply(compute_technology_today, axis=1)
    df["Technology_Calculated"] = np.where(df["Technology_Today"] != "UNKNOWN", "Calculated", "Manual")

    # Merge CI settings
    df = left_merge(df, tabCIsettings, ["PLANNER_UNIQUE_IDENTIFIER"], ["PLANNER_UNIQUE_IDENTIFIER"])

    # Fill UNKNOWN with manual data
    if "Location Manually" in df.columns:
        df["Location_Today"] = np.where(
            (df["Location_Today"] == "UNKNOWN")
            & df["Location Manually"].notna()
            & (df["Location Manually"].astype(str).str.strip() != ""),
            df["Location Manually"],
            df["Location_Today"],
        )

    if "Technology Manually" in df.columns:
        df["Technology_Today"] = np.where(
            (df["Technology_Today"] == "UNKNOWN")
            & df["Technology Manually"].notna()
            & (df["Technology Manually"].astype(str).str.strip() != ""),
            df["Technology Manually"],
            df["Technology_Today"],
        )

    # Replace unknown CI Summary
    if "CI Summary" in df.columns:
        df["CI Summary"] = df["CI Summary"].fillna("Unkown CI Summary")
    else:
        df["CI Summary"] = "Unkown CI Summary"

    # Concats
    df["S2T2T"] = (
        df["CI Summary"].astype(str)
        + "_"
        + df["DISPOSITION_OPTION_STEP_1"].astype(str)
        + "_"
        + df["DISPOSITION_OPTION_STEP_2"].astype(str)
    )
    df["S2T-1"] = df["DISPOSITION_OPTION_STEP_1"].astype(str)
    df["S2T-2"] = df["DISPOSITION_OPTION_STEP_1"].astype(str) + "_" + df["DISPOSITION_OPTION_STEP_2"].astype(str)

    # Step0 merge: PQ uses tabStepsettings, keyed on Technology Today.
    # Your main table uses Technology_Today, your settings uses "Technology Today" -> map to a common name.
    if "Technology Today" in tabStep0settings.columns and "Technology_Today" not in tabStep0settings.columns:
        tabStep0settings = tabStep0settings.rename(columns={"Technology Today": "Technology_Today"})
    df = left_merge(df, tabStep0settings, ["Technology_Today"], ["Technology_Today"])

    # Step1/Step2 merges
    df = left_merge(df, tabStep1settings, ["DISPOSITION_OPTION_STEP_1"], ["DISPOSITION_OPTION_STEP_1"])
    df = left_merge(df, tabStep2settings, ["DISPOSITION_OPTION_STEP_2"], ["DISPOSITION_OPTION_STEP_2"])

    # S2T merges
    df = left_merge(df, tabS2T1settings, ["S2T-1"], ["S2T1"])
    df = left_merge(df, tabS2T2settings, ["S2T-2"], ["S2T2"])
    df = left_merge(df, tabS2T2Tsettings, ["S2T2T"], ["S2T2T"])

    # NAR baseline
    df = left_merge(df, NAR_ReportBaseLine, ["NAR_INSTANCE_ID"], ["NAR ID"])
    if "Certified Decom Candidate" in df.columns:
        df = df.rename(columns={"Certified Decom Candidate": "NAR App Status Decom"})
    if "Instance Planned Retirement Date" in df.columns:
        df = df.rename(columns={"Instance Planned Retirement Date": "NAR App Planned Retirement Date"})

    # EAP consumers
    df = left_merge(df, tabEAP_Grid_Consumers, ["NAR_INSTANCE_ID"], ["NAR ID"])

    # Platform NARs
    df = left_merge(df, tabPlatformNARs, ["NAR_INSTANCE_ID"], ["NAR ID"])

    # Remove helper columns
    for c in ["Location Manually", "Technology Manually"]:
        if c in df.columns:
            df = df.drop(columns=[c])

    # Chase decision columns
    def chase_decision(dispo_col: str, chase_flag_col: str) -> pd.Series:
        ensure_col(df, "ChaserBlocker", "")
        ensure_col(df, chase_flag_col, "")
        ensure_col(df, dispo_col, "")

        return np.select(
            [
                df["ChaserBlocker"].astype(str) == "YES",
                df["CI Summary"].astype(str) == "No response required",
                df[chase_flag_col].astype(str) == "NO",
                df[dispo_col].notna()
                & (df[dispo_col].astype(str).str.strip() != "")
                & (df[dispo_col].astype(str) != "0_EMPTY"),
            ],
            [
                "not required (Chaser Block)",
                "not required (Non-relevant CI)",
                "not required",
                "filled",
            ],
            default="pending",
        )

    df["Step1-Dispo Chase Decision"] = chase_decision("DISPOSITION_OPTION_STEP_1", "S2T1-CIO to Dispo Chase")
    df["Step2-Dispo Chase Decision"] = chase_decision("DISPOSITION_OPTION_STEP_2", "S2T2-CIO to Dispo Chase")
    df["Step1-Time Chase Decision"] = chase_decision("DISPOSITION_TIMELINE_STEP_1", "S2T1-CIO to Time Chase")
    df["Step2-Time Chase Decision"] = chase_decision("DISPOSITION_TIMELINE_STEP_2", "S2T2-CIO to Time Chase")

    # Schedules
    df = left_merge(df, tabSchedule_V2V, ["SERVER_NAME"], ["Full_Server_Name"])
    df = left_merge(df, tabSchedule_P2V, ["SERVER_NAME"], ["SERVER_NAME"])
    df = left_merge(df, tabSchedule_P2P, ["SERVER_NAME"], ["SERVER_NAME"])

    # Part of Migration Initiative (P2V, P2P, V2V) - robust fix (no NA ambiguity)
    ensure_col(df, "V2V-Scope", "")
    ensure_col(df, "P2P Scope", "")
    v2v = df["V2V-Scope"].astype("string").fillna("").str.strip()
    p2p = df["P2P Scope"].astype("string").fillna("").str.strip()
    mask = v2v.ne("") | p2p.ne("")
    df["Part of Migration Initiative (P2V, P2P, V2V)"] = ""
    df.loc[mask, "Part of Migration Initiative (P2V, P2P, V2V)"] = "Yes"

    # Replace EMPTY -> "No" for these
    df = fill_empty_cells(
        df,
        cols=["EAP/Grid Consumer", "Platform_Provider", "Baseline Aug", "Baseline Sept", "Baseline Oct", "Baseline Nov", "Baseline Dec"],
        token="No",
    )

    # Remove columns (your big drop list)
    cols_to_drop = [
        "IG_ROW_UPDATE_ALLOWED","ESTATE_MANAGEMENT_SCOPE","REPORTING_GROUP","SCHEDULING_RECID","LINE_OF_BUSINESS",
        "PLANNING_ID","TREATMENT","TARGET_ACTUAL","COMMENTS","PEAKOFPEAKSCPUUSAGE","AVGCPUUSAGE","CORES",
        "PEAKOFPEAKSMEMPERCENTAGE","AVGMEMPERCENTAGE","MEMORY_GB","TARGET_INFRA_REQUEST_DATE","TARGET_INFRA_DELIVERY_DATE",
        "TARGET_INFRA_CUTOVER_DATE","TARGET_INFRA_DECOM_DATE","TARGET_DECOM_DATE_BASELINE","BOW_YEAR","REPLACEMENT_RFS_NO",
        "DECOM_RFS_NO","LIFECYCLE_STATUS","ADJUSTED_CORES","HW_EOL_YEAR","OS_EOL_YEAR","DATABASE_EOL_YEAR",
        "SUPPORT_GROUP","DATABASE_INSTANCE_COUNT","SERVER_VIRTUAL_COUNT","APPLICATIONCRITICALITYCLASS","APP_PLANNED_RETIRE_DATE",
        "APP_PLAN_RETIRE_DATE_CERTIFIED","APP_INST_INVESTMENT_STRATEGY","APP_RECOVERY_CLASS","APP_TECHNOLOGY_RTO",
        "MAS_INSCOPE","MAS_CRITICAL","REG_CRITICAL","TRC_REG_OR_CRITICAL","LATEST_MONTHLY_COST","APPTIO_ASOF",
        "MIGRATION_STATUS","TRACKER_MODIFIED_BY","TRACKER_MODIFIED_DATE","REMOVAL_DATE","REMOVED_REASON",
        "EFFECTIVE_CLASSIFICATION","DECOM_TARGET","DLINK_TICKET_NUMBER","DLINK_SUBMITTED_DATE","DLINK_CLOSED_DATE","DLINK_STATE",
        "DLINK_STAGE","EM_FILTER_TAG","DECOM_DATE_IS_PLACEHOLDER","INFRA_RECEIVED_FLAG","RELATED_ORDERS",
        "SERVER_INSTALL_DATE","REBUILD_DATE","SOONEST_OBSOLESCENCE_DATE","TR_OBSOLESCENCE_DATE_HW","TRC_OS_OBSOLESCENCE_DATE",
        "TRC_APP_TR_COMPLIANT","TRC_CI_TR_COMPLIANT","LEGAL_HOLD_CODE","RECORDS_MGMT_CODE","ARCHIVE_CERTIFICATION_CODE",
        "CLOUD_APP_DELINK_DATE","CIRRUS_R_TYPE","SERVICE_URL","HOSTING_CLUSTER_NAMES","JIRA_TEXT","ONEMI_PRODUCT_INSTANCE_ID",
        "CWB_TYPE","CWB_DUE_DATE","CWB_COMMITMENT_DATE","OTR_DECOM_TARGET_DATE","EM_INITIATIVES_1","HSF_REF","OCP_VERSION",
        "OTR_MIGRATION_STATUS","VERITAS_CI","VENDOR_LICENSED_APPLICATION","VENDOR_NAME","REMOVED_CI_CIRRUS_SCOPE","CLUSTER_ID",
        "CLUSTER_NAME","CIO","CI_ID","SERVER_ID","PARENT_SERVER","DATABASE_ID","REMOVED_FLAG","APP_CI_REL_CREATED_DATE",
        "RE_INSTATEMENT_DATE","ATC_ACTION_2026","Project Flag","HSF_COMMITTED_DATE","ATC_ACTION_2025",
        "S2T1-CIO to Dispo Chase","S2T2-CIO to Dispo Chase","Today-L1 Grouping","Step1-L1 Grouping","Step2-L1 Grouping",
        "S2T1-L1 Grouping","S2T2-L1 Grouping","INTERIM_TARGET_PRODUCT_2025_2026","FINAL_TARGET_PRODUCT_2028",
        "Underpinning_Server_CIs"
    ]
    df = df.drop(columns=[c for c in cols_to_drop if c in df.columns], errors="ignore")

    # Data Residency
    df = left_merge(df, tabData_Residency, ["NAR_INSTANCE_ID"], ["NAR-ID"])

    # V2V Sub Scope
    df["V2V Sub Scope"] = df.apply(compute_v2v_sub_scope, axis=1)

    # Vendor_Today
    df["Vendor_Today"] = df.apply(compute_vendor_today, axis=1)

    # Shared/Dedicate tagging
    df = left_merge(df, Server_Shared_Dedicate_Tagging, ["SERVER_NAME"], ["SERVER_NAME"])
    if "SharedDedicated_Server" in df.columns and "Shared\\Dedicate_Server" not in df.columns:
        df = df.rename(columns={"SharedDedicated_Server": "Shared\\Dedicate_Server"})

    # Location_Today2 on HP_DC_NAME
    df = left_merge(df, Location_Today2, ["HP_DC_NAME"], ["HP_DC_NAME"])
    if "Location_Today" in Location_Today2.columns:
        df = df.rename(columns={"Location_Today": "Location_Today_DC-Name-Based"})

    # Final reorder (keep what exists)
    final_order = [
        "PLANNER_UNIQUE_IDENTIFIER","CI_CATEGORY","PRODUCT_PORTFOLIO_NAR","IS_ACTUAL_DATABASE_CI","REPORTING_UNIT",
        "NAR_INSTANCE_ID","INSTANCE_NAME","CIO_MINUS_1","PORTFOLIO_OWNER","PORTFOLIO_OWNER_DELEGATE",
        "INSTANCEITA0","INSTANCEITA0_DELEGATE","CI_NAME","REGION","COUNTRY","BUILDING","HP_DC_NAME","CITY",
        "SERVER_NAME","PARENT_SERVER_NAME","CLASSIFICATION","SERVER_TYPE","HOST_TYPE","MODEL","OPERATING_SYSTEM",
        "IN_DMZ","KYNDRYL_CATEGORY","DATABASE_TYPE","DATABASE_NAME","DB_NAME","DB_SUBCATEGORY","DATABASE_VERSION",
        "HP_P4_P7_BILLABLE","GTI_INITIATIVES_2026","APPLICATION_NAME","HP_DB_SUPPORTED","P4_P7_REASON",
        "S2T2T","S2T-1","S2T-2","CI Summary",
        "Location_Today","Location_Calculated","Location_Today_DC-Name-Based","OneMI_LOCATION_TODAY",
        "Technology_Today","Technology_Calculated","OneMI_TECHNOLOGY_TODAY",
        "Vendor_Today","OneMI_VENDOR_TODAY",
        "Today Platform","CI_REQUIRING_PLANS","KPI_BASELINE","KPI_AND_PLANS","EAP/Grid Consumer",
        "Baseline Aug","Baseline Sept","Baseline Oct","Baseline Nov","Baseline Dec",
        "DISPOSITION_OPTION_STEP_1","DISPOSITION_TIMELINE_STEP_1","DISPOSITION_COMMENTS_STEP_1",
        "DISPOSITION_OPTION_STEP_2","DISPOSITION_TIMELINE_STEP_2","DISPOSITION_COMMENTS_STEP_2",
        "GTI_INITIATIVES_2025","Anomalie YN","AnomalieBlocker","ChaserBlocker","NAR_DECOM_ANOMALY",
        "S2T1-CIO to Time Chase","S2T1-Project / Same S2T","S2T2-CIO to Time Chase","S2T2-Project / Same S2T",
        "CIO Chase YN","Step1-Dispo Chase Decision","Step2-Dispo Chase Decision","Step1-Time Chase Decision",
        "Step2-Time Chase Decision","NAR App Status Decom","NAR App Planned Retirement Date",
        "CONSOLIDATED_DECOM_DATE","CONSOLIDATED_DECOM_DATE_SOURCE","SERVICE","SERVICE_OFFERING","PAAS",
        "S2T2T-Grouped","Today-L0 Grouping","End Step 1 Platform","Step1-L0 Grouping","End Step 2 Platform",
        "Step2-L0 Grouping","S2T1-L0 Grouping","S2T2-L0 Grouping",
        "TRC_HW_PLAN_TYPE","TRC_HW_EARLIEST_REMEDIATION_DATE","TRC_SW_PLAN_TYPE","Platform_Provider",
        "TRC_SW_EARLIEST_REMEDIATION_DATE","INITIATIVES","PLANNED_MIGRATION_DATE","SOURCE_MAPPING",
        "Part of Migration Initiative (P2V, P2P, V2V)","V2V-Scope","V2V Sub Scope",
        "V2V_Migration_Wave_Start_Date","V2V_Migration_Wave_End_Date","P2V Scope","P2V Migration Plan",
        "P2P Scope","P2P Migration Plan","Data_Residency","Shared\\Dedicate_Server"
    ]
    existing = [c for c in final_order if c in df.columns]
    remaining = [c for c in df.columns if c not in existing]
    df = df[existing + remaining]

    return df


def main() -> None:
    df_out = run_pipeline()
    Path(os.path.dirname(OUTPUT_PATH)).mkdir(parents=True, exist_ok=True)
    df_out.to_excel(OUTPUT_PATH, index=False)
    print(f"Done. Rows: {len(df_out):,}. Output written to: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
